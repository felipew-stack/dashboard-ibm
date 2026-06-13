import streamlit as st
import pandas as pd
import plotly.express as px
import base64
import gspread
from io import BytesIO
from datetime import datetime
from google.oauth2.service_account import Credentials
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Dashboard IBM", layout="wide")

SHEET_ID = "1gNBenj4s19pOtlNbIAZp0_CYpAXBidxXbAtg9hdOXcM"

ABA_TOTAL = "Localidades Total"
ABA_PENDENTES = "Localidades Pendentes"
ABA_CONCLUIDAS = "Localidades Concluídas"
ABA_BUSCAS = "Buscas por Agente"

ARQUIVO_LOGO = "logo_3am.png"

COLUNAS_TOTAL = ["UF", "Cidade"]
COLUNAS_CONCLUIDAS = ["UF", "Cidade"]

COLUNAS_PENDENTES = [
    "UF",
    "Cidade",
    "Data da Solicitação",
    "Previsão",
    "Prioridade",
    "Relatório Detalhado"
]

COLUNAS_BUSCAS = [
    "Agente",
    "UF",
    "Cidade",
    "Data da Busca",
    "Status",
    "Observação"
]

AGENTES = ["Felipe", "Diovane"]

STATUS_BUSCA = [
    "Em busca",
    "Contato realizado",
    "Aguardando retorno",
    "Recurso encontrado",
    "Sem retorno",
    "Finalizado"
]


def conectar_google_sheets():
    try:
        escopos = [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive"
        ]

        credenciais = Credentials.from_service_account_info(
            st.secrets["gcp_service_account"],
            scopes=escopos
        )

        cliente = gspread.authorize(credenciais)
        return cliente.open_by_key(SHEET_ID)

    except Exception as erro:
        st.error("Erro ao conectar com o Google Sheets.")
        st.write(erro)
        st.stop()


def carregar_aba(planilha, nome_aba, colunas):
    try:
        aba = planilha.worksheet(nome_aba)
        dados = aba.get_all_records()
        df = pd.DataFrame(dados)

        df.columns = [str(coluna).strip() for coluna in df.columns]

        for coluna in colunas:
            if coluna not in df.columns:
                df[coluna] = ""

        df = df[colunas]
        return df

    except gspread.WorksheetNotFound:
        aba = planilha.add_worksheet(
            title=nome_aba,
            rows=100,
            cols=len(colunas)
        )
        aba.update([colunas])
        return pd.DataFrame(columns=colunas)

    except Exception as erro:
        st.error(f"Erro ao carregar a aba '{nome_aba}'.")
        st.write(erro)
        st.stop()


def salvar_aba(planilha, nome_aba, df, colunas, permitir_vazio=False):
    try:
        if df is None:
            st.error(
                f"Salvamento bloqueado: dados da aba '{nome_aba}' não encontrados."
            )
            st.stop()

        if df.empty and not permitir_vazio:
            st.error(
                f"Salvamento bloqueado: a aba '{nome_aba}' está vazia. "
                "Nada foi alterado no Google Sheets."
            )
            st.stop()

        aba = planilha.worksheet(nome_aba)

        df_salvar = df.copy()
        df_salvar.columns = [str(coluna).strip()
                             for coluna in df_salvar.columns]

        for coluna in colunas:
            if coluna not in df_salvar.columns:
                df_salvar[coluna] = ""

        df_salvar = df_salvar[colunas].fillna("")
        valores = [df_salvar.columns.tolist()] + \
            df_salvar.astype(str).values.tolist()

        aba.clear()
        aba.update(valores)

    except gspread.WorksheetNotFound:
        aba = planilha.add_worksheet(
            title=nome_aba,
            rows=100,
            cols=len(colunas)
        )

        df_salvar = df.copy()

        for coluna in colunas:
            if coluna not in df_salvar.columns:
                df_salvar[coluna] = ""

        df_salvar = df_salvar[colunas].fillna("")
        valores = [df_salvar.columns.tolist()] + \
            df_salvar.astype(str).values.tolist()
        aba.update(valores)

    except Exception as erro:
        st.error(f"Erro ao salvar a aba '{nome_aba}'.")
        st.write(erro)
        st.stop()


def ordenar_localidades(df, colunas):
    df = df.copy()
    df.columns = [str(coluna).strip() for coluna in df.columns]

    for coluna in colunas:
        if coluna not in df.columns:
            df[coluna] = ""

    df = df[colunas]

    df["UF"] = df["UF"].astype(str).str.strip().str.upper()
    df["Cidade"] = df["Cidade"].astype(str).str.strip()

    df = df[
        (df["UF"] != "") &
        (df["UF"].str.lower() != "nan") &
        (df["Cidade"] != "") &
        (df["Cidade"].str.lower() != "nan")
    ]

    df = df.sort_values(by=["Cidade", "UF"]).reset_index(drop=True)
    return df


def organizar_pendentes(df):
    df = df.copy()
    df.columns = [str(coluna).strip() for coluna in df.columns]

    for coluna in COLUNAS_PENDENTES:
        if coluna not in df.columns:
            df[coluna] = ""

    df = df[COLUNAS_PENDENTES]

    df["UF"] = df["UF"].astype(str).str.strip().str.upper()
    df["Cidade"] = df["Cidade"].astype(str).str.strip()
    df["Prioridade"] = df["Prioridade"].astype(str).str.strip()

    df["Data da Solicitação"] = df["Data da Solicitação"].astype(str).replace(
        ["nan", "NaT", "None"], ""
    )

    df["Previsão"] = df["Previsão"].astype(str).replace(
        ["nan", "NaT", "None"], ""
    )

    df = df[df["Cidade"] != ""].reset_index(drop=True)
    return df


def organizar_buscas(df):
    df = df.copy()
    df.columns = [str(coluna).strip() for coluna in df.columns]

    for coluna in COLUNAS_BUSCAS:
        if coluna not in df.columns:
            df[coluna] = ""

    df = df[COLUNAS_BUSCAS]

    df["Agente"] = df["Agente"].astype(str).str.strip()
    df["UF"] = df["UF"].astype(str).str.strip().str.upper()
    df["Cidade"] = df["Cidade"].astype(str).str.strip()
    df["Data da Busca"] = df["Data da Busca"].astype(str).replace(
        ["nan", "NaT", "None"], ""
    )
    df["Status"] = df["Status"].astype(str).str.strip()
    df["Observação"] = df["Observação"].astype(str).replace(
        ["nan", "NaT", "None"], ""
    )

    df = df[df["Cidade"] != ""].reset_index(drop=True)
    return df


def salvar_google_sheets(df_total, df_pendentes, df_concluidas, df_buscas):
    planilha = conectar_google_sheets()

    df_total = ordenar_localidades(df_total, COLUNAS_TOTAL)
    df_pendentes = organizar_pendentes(df_pendentes)
    df_concluidas = ordenar_localidades(df_concluidas, COLUNAS_CONCLUIDAS)
    df_buscas = organizar_buscas(df_buscas)

    salvar_aba(planilha, ABA_TOTAL, df_total, COLUNAS_TOTAL)
    salvar_aba(planilha, ABA_PENDENTES, df_pendentes,
               COLUNAS_PENDENTES, permitir_vazio=True)
    salvar_aba(planilha, ABA_CONCLUIDAS, df_concluidas,
               COLUNAS_CONCLUIDAS, permitir_vazio=True)
    salvar_aba(planilha, ABA_BUSCAS, df_buscas,
               COLUNAS_BUSCAS, permitir_vazio=True)


def imagem_base64(caminho):
    try:
        with open(caminho, "rb") as arquivo:
            return base64.b64encode(arquivo.read()).decode()
    except FileNotFoundError:
        return None


def corrigir_relatorio(texto):
    if pd.isna(texto):
        return ""

    texto = str(texto).strip()

    correcoes = {
        "Encontramos recursos na localidades": "Encontramos recursos na localidade",
        "Encontramos recursos nas localidades": "Encontramos recursos na localidade",
        "Encontramos recursos no localidades": "Encontramos recursos na localidade",
        "Encontramos recursos na localidade, mas declinaram": "Encontramos recursos na localidade, mas eles declinaram",
        "porem": "porém",
        "Porem": "Porém",
        "tecnicos": "técnicos",
        "Tecnicos": "Técnicos",
        "nao": "não",
        "Nao": "Não",
        "deu retorno": "retornou",
        "nega a proposta": "negou a proposta",
        "negou a proprosta": "negou a proposta",
        "proprosta": "proposta"
    }

    for errado, certo in correcoes.items():
        texto = texto.replace(errado, certo)

    return texto


RELATORIOS_PADRAO = [
    "Em busca de recurso técnico para a localidade. Até o momento, os profissionais acionados não retornaram ou não demonstraram disponibilidade para atendimento.",
    "Seguimos realizando buscas na região. Foram identificados alguns recursos, porém ainda não obtivemos confirmação para fechamento da parceria.",
    "A localidade permanece em aberto. Os contatos realizados até o momento não resultaram em profissionais disponíveis para atendimento.",
    "Foram efetuadas buscas e contatos com recursos da região, porém ainda não foi possível validar um profissional para cobertura da demanda.",
    "Seguimos priorizando a localidade. Os profissionais encontrados até o momento declinaram da oportunidade ou não responderam às tentativas de contato.",
    "A equipe continua em prospecção de recursos técnicos. Até o momento não foi possível concluir a validação de um profissional apto para atendimento.",
    "Foram realizadas novas tentativas de contato na região. Permanecemos aguardando retorno dos profissionais acionados.",
    "A localidade segue em acompanhamento. As buscas continuam sendo realizadas para identificação de recursos compatíveis com a necessidade do projeto.",
    "Identificamos possíveis recursos na região, porém ainda não houve aceite para execução dos atendimentos solicitados.",
    "Seguimos com a prospecção de profissionais para a localidade. Até o momento não obtivemos sucesso na confirmação de disponibilidade dos recursos contatados."
]


def preencher_relatorios_vazios(df):
    df = df.copy()

    if "Relatório Detalhado" not in df.columns:
        df["Relatório Detalhado"] = ""

    for indice in df.index:
        relatorio = str(df.loc[indice, "Relatório Detalhado"]).strip()

        if relatorio in ["", "nan", "None", "NaT"]:
            df.loc[indice, "Relatório Detalhado"] = RELATORIOS_PADRAO[
                indice % len(RELATORIOS_PADRAO)
            ]
        else:
            df.loc[indice, "Relatório Detalhado"] = corrigir_relatorio(
                relatorio)

    return df


def gerar_excel_relatorio(df):
    output = BytesIO()
    df_export = df.copy()

    colunas_existentes = [
        coluna for coluna in COLUNAS_PENDENTES
        if coluna in df_export.columns
    ]

    df_export = df_export[colunas_existentes]

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_export.to_excel(writer, index=False,
                           sheet_name="Relatório Detalhado")
        worksheet = writer.sheets["Relatório Detalhado"]

        azul_escuro = "003B71"
        branco = "FFFFFF"
        cinza_claro = "F2F6FA"
        vermelho = "F8D7DA"
        amarelo = "FFF3CD"
        verde = "D4EDDA"

        borda_fina = Side(style="thin", color="D9E2EC")

        for cell in worksheet[1]:
            cell.fill = PatternFill(
                start_color=azul_escuro,
                end_color=azul_escuro,
                fill_type="solid"
            )
            cell.font = Font(color=branco, bold=True)
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )
            cell.border = Border(
                left=borda_fina,
                right=borda_fina,
                top=borda_fina,
                bottom=borda_fina
            )

        worksheet.row_dimensions[1].height = 28

        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = Border(
                    left=borda_fina,
                    right=borda_fina,
                    top=borda_fina,
                    bottom=borda_fina
                )

                if cell.row % 2 == 0:
                    cell.fill = PatternFill(
                        start_color=cinza_claro,
                        end_color=cinza_claro,
                        fill_type="solid"
                    )

        if "Prioridade" in df_export.columns:
            coluna_prioridade = df_export.columns.get_loc("Prioridade") + 1

            for row in range(2, worksheet.max_row + 1):
                celula = worksheet.cell(row=row, column=coluna_prioridade)
                valor = str(celula.value).strip()

                if valor == "Alta":
                    celula.fill = PatternFill(
                        start_color=vermelho,
                        end_color=vermelho,
                        fill_type="solid"
                    )
                    celula.font = Font(bold=True)

                elif valor == "Média":
                    celula.fill = PatternFill(
                        start_color=amarelo,
                        end_color=amarelo,
                        fill_type="solid"
                    )
                    celula.font = Font(bold=True)

                elif valor == "Baixa":
                    celula.fill = PatternFill(
                        start_color=verde,
                        end_color=verde,
                        fill_type="solid"
                    )
                    celula.font = Font(bold=True)

        larguras = {
            "UF": 10,
            "Cidade": 28,
            "Data da Solicitação": 22,
            "Previsão": 18,
            "Prioridade": 16,
            "Relatório Detalhado": 90
        }

        for idx, coluna in enumerate(df_export.columns, start=1):
            letra = get_column_letter(idx)
            worksheet.column_dimensions[letra].width = larguras.get(coluna, 25)

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

    return output.getvalue()


planilha = conectar_google_sheets()

df_total = carregar_aba(planilha, ABA_TOTAL, COLUNAS_TOTAL)
df_pendentes = carregar_aba(planilha, ABA_PENDENTES, COLUNAS_PENDENTES)
df_concluidas = carregar_aba(planilha, ABA_CONCLUIDAS, COLUNAS_CONCLUIDAS)
df_buscas = carregar_aba(planilha, ABA_BUSCAS, COLUNAS_BUSCAS)

df_total = ordenar_localidades(df_total, COLUNAS_TOTAL)
df_pendentes = organizar_pendentes(df_pendentes)
df_concluidas = ordenar_localidades(df_concluidas, COLUNAS_CONCLUIDAS)
df_buscas = organizar_buscas(df_buscas)

df_pendentes = preencher_relatorios_vazios(df_pendentes)

total = len(df_total)
pendentes = len(df_pendentes)
concluidas = len(df_concluidas)
percentual = round((concluidas / total) * 100, 1) if total > 0 else 0

alta = len(df_pendentes[df_pendentes["Prioridade"] == "Alta"])
media = len(df_pendentes[df_pendentes["Prioridade"] == "Média"])
baixa = len(df_pendentes[df_pendentes["Prioridade"] == "Baixa"])

st.markdown(
    """
    <style>
        .logo-box {
            background-color: #FFFFFF;
            padding: 14px 20px;
            border-radius: 16px;
            display: inline-flex;
            align-items: center;
            justify-content: center;
            box-shadow: 0px 2px 10px rgba(0,0,0,0.12);
            border: 1px solid rgba(0,0,0,0.08);
        }

        .logo-box img {
            width: 230px;
        }

        .header-title {
            padding-top: 12px;
        }
    </style>
    """,
    unsafe_allow_html=True
)

logo_col, titulo_col = st.columns([1.2, 4])

with logo_col:
    logo_base64 = imagem_base64(ARQUIVO_LOGO)

    if logo_base64:
        st.markdown(
            f"""
            <div class="logo-box">
                <img src="data:image/png;base64,{logo_base64}">
            </div>
            """,
            unsafe_allow_html=True
        )
    else:
        st.warning(f"Logo '{ARQUIVO_LOGO}' não encontrada.")

with titulo_col:
    st.markdown('<div class="header-title">', unsafe_allow_html=True)
    st.title("Projeto IBM")
    st.caption(
        "Monitoramento de localidades abertas, pendentes, concluídas e relatório detalhado"
    )
    st.markdown("</div>", unsafe_allow_html=True)

aba_dashboard, aba_relatorio, aba_buscas, aba_total, aba_concluidas = st.tabs([
    "Dashboard",
    "Relatório Detalhado",
    "Gestão de Buscas",
    "Localidades Total",
    "Localidades Concluídas"
])
with aba_dashboard:
    st.divider()

    col1, col2, col3, col4 = st.columns(4)

    col1.metric("Total de Localidades", total)
    col2.metric("Localidades Concluídas", concluidas)
    col3.metric("Localidades Pendentes", pendentes)
    col4.metric("Percentual de Conclusão", f"{percentual}%")

    st.progress(percentual / 100)

    col5, col6, col7 = st.columns(3)

    col5.metric("🔴 Prioridade Alta", alta)
    col6.metric("🟡 Prioridade Média", media)
    col7.metric("🟢 Prioridade Baixa", baixa)

    st.divider()

    st.subheader("Adicionar Nova Localidade")
    st.caption(
        "Cadastre uma nova localidade. Ela será adicionada em Localidades Total e também em Localidades Pendentes."
    )

    with st.expander("Adicionar nova localidade", expanded=False):
        add_col1, add_col2, add_col3 = st.columns(3)

        with add_col1:
            nova_uf = st.text_input("UF", max_chars=2).upper()
            nova_cidade = st.text_input("Cidade")

        with add_col2:
            nova_data = st.date_input("Data da Solicitação")
            nova_previsao = st.date_input("Previsão")

        with add_col3:
            nova_prioridade = st.selectbox(
                "Prioridade",
                ["Alta", "Média", "Baixa"]
            )

            novo_relatorio = st.text_area(
                "Relatório Detalhado",
                height=120,
                placeholder="Digite a situação da localidade..."
            )

        if st.button("Adicionar Localidade"):
            if not nova_uf or not nova_cidade:
                st.warning("Preencha pelo menos UF e Cidade.")
            else:
                cidade_formatada = nova_cidade.strip().title()
                uf_formatada = nova_uf.strip().upper()

                existe_total = df_total[
                    (df_total["Cidade"].astype(str).str.upper() == cidade_formatada.upper()) &
                    (df_total["UF"].astype(str).str.upper()
                     == uf_formatada.upper())
                ]

                existe_pendente = df_pendentes[
                    (df_pendentes["Cidade"].astype(str).str.upper() == cidade_formatada.upper()) &
                    (df_pendentes["UF"].astype(
                        str).str.upper() == uf_formatada.upper())
                ]

                if not existe_total.empty:
                    st.warning(
                        "Essa localidade já existe em Localidades Total.")
                elif not existe_pendente.empty:
                    st.warning(
                        "Essa localidade já existe em Localidades Pendentes.")
                else:
                    nova_linha_total = {
                        "UF": uf_formatada,
                        "Cidade": cidade_formatada
                    }

                    nova_linha_pendente = {
                        "UF": uf_formatada,
                        "Cidade": cidade_formatada,
                        "Data da Solicitação": nova_data.strftime("%d/%m/%Y"),
                        "Previsão": nova_previsao.strftime("%d/%m/%Y"),
                        "Prioridade": nova_prioridade,
                        "Relatório Detalhado": novo_relatorio.strip()
                    }

                    if not nova_linha_pendente["Relatório Detalhado"]:
                        nova_linha_pendente["Relatório Detalhado"] = RELATORIOS_PADRAO[
                            len(df_pendentes) % len(RELATORIOS_PADRAO)
                        ]

                    df_total = pd.concat(
                        [df_total, pd.DataFrame([nova_linha_total])],
                        ignore_index=True
                    )

                    df_pendentes = pd.concat(
                        [df_pendentes, pd.DataFrame([nova_linha_pendente])],
                        ignore_index=True
                    )

                    salvar_google_sheets(
                        df_total, df_pendentes, df_concluidas, df_buscas)

                    st.success("Localidade adicionada com sucesso.")
                    st.rerun()

    st.divider()

    graf1, graf2 = st.columns(2)

    with graf1:
        dados_status = pd.DataFrame({
            "Status": ["Concluídas", "Pendentes"],
            "Quantidade": [concluidas, pendentes]
        })

        fig_status = px.pie(
            dados_status,
            names="Status",
            values="Quantidade",
            hole=0.55,
            title="Status Geral das Localidades",
            color="Status",
            color_discrete_map={
                "Concluídas": "#00A6D6",
                "Pendentes": "#003B71"
            }
        )

        fig_status.update_traces(
            textposition="inside",
            textinfo="percent+label"
        )

        st.plotly_chart(fig_status, use_container_width=True)

    with graf2:
        uf_count = df_pendentes["UF"].value_counts().reset_index()
        uf_count.columns = ["UF", "Quantidade"]

        fig_uf = px.bar(
            uf_count,
            x="UF",
            y="Quantidade",
            text_auto=True,
            title="Pendências por Estado",
            color="Quantidade",
            color_continuous_scale="Blues"
        )

        fig_uf.update_layout(
            xaxis_title="Estado",
            yaxis_title="Quantidade de Pendências",
            showlegend=False
        )

        st.plotly_chart(fig_uf, use_container_width=True)

    st.divider()

    st.subheader("Localidades por Estado")
    st.caption(
        "Clique em um estado para listar as localidades pendentes. Depois clique na localidade para ver o relatório detalhado ali mesmo."
    )

    if "uf_expandida" not in st.session_state:
        st.session_state["uf_expandida"] = None

    if df_pendentes.empty:
        st.info("Não há localidades pendentes para exibir.")
    else:
        total_colunas_uf = min(5, len(uf_count)) if len(uf_count) > 0 else 1
        colunas_uf = st.columns(total_colunas_uf)

        for posicao, (_, linha_uf) in enumerate(uf_count.iterrows()):
            uf_atual = str(linha_uf["UF"]).strip()
            quantidade_uf = int(linha_uf["Quantidade"])

            with colunas_uf[posicao % total_colunas_uf]:
                if st.button(
                    f"{uf_atual} ({quantidade_uf})",
                    key=f"botao_uf_{uf_atual}"
                ):
                    if st.session_state["uf_expandida"] == uf_atual:
                        st.session_state["uf_expandida"] = None
                    else:
                        st.session_state["uf_expandida"] = uf_atual

        if st.session_state["uf_expandida"]:
            uf_selecionada = st.session_state["uf_expandida"]

            localidades_estado = df_pendentes[
                df_pendentes["UF"] == uf_selecionada
            ].copy()

            st.markdown(f"### Localidades Pendentes - {uf_selecionada}")

            for _, linha_localidade in localidades_estado.iterrows():
                cidade = str(linha_localidade.get("Cidade", "")).strip()
                prioridade_localidade = str(
                    linha_localidade.get("Prioridade", "")
                ).strip()
                data_solicitacao = linha_localidade.get(
                    "Data da Solicitação",
                    "-"
                )
                previsao = linha_localidade.get("Previsão", "-")
                relatorio = linha_localidade.get("Relatório Detalhado", "")

                with st.expander(
                    f"📍 {cidade} | Prioridade: {prioridade_localidade}",
                    expanded=False
                ):
                    st.markdown(f"**UF:** {uf_selecionada}")
                    st.markdown(f"**Cidade:** {cidade}")
                    st.markdown(f"**Prioridade:** {prioridade_localidade}")
                    st.markdown(f"**Data da Solicitação:** {data_solicitacao}")
                    st.markdown(f"**Previsão:** {previsao}")
                    st.markdown("**Relatório Detalhado:**")
                    st.write(str(relatorio))

    st.divider()

    st.subheader("Resumo das Localidades Pendentes")
    st.caption(
        "Consulte, filtre, altere a prioridade e marque localidades como concluídas diretamente pelo Dashboard."
    )

    filtro1, filtro2, filtro3 = st.columns(3)

    with filtro1:
        busca = st.text_input("Pesquisar cidade")

    with filtro2:
        prioridade = st.selectbox(
            "Filtrar por prioridade",
            ["Todas", "Alta", "Média", "Baixa"]
        )

    with filtro3:
        uf = st.selectbox(
            "Filtrar por UF",
            ["Todas"] + sorted(df_pendentes["UF"].dropna().unique().tolist())
        )

    tabela = df_pendentes.copy()
    tabela["__linha_original"] = tabela.index
    tabela["Concluir"] = False

    colunas_editor = [
        "UF",
        "Cidade",
        "Data da Solicitação",
        "Previsão",
        "Prioridade",
        "Concluir",
        "__linha_original"
    ]

    tabela = tabela[colunas_editor]

    if busca:
        tabela = tabela[
            tabela["Cidade"].astype(str).str.contains(
                busca,
                case=False,
                na=False
            )
        ]

    if prioridade != "Todas":
        tabela = tabela[tabela["Prioridade"] == prioridade]

    if uf != "Todas":
        tabela = tabela[tabela["UF"] == uf]

    tabela_editada = st.data_editor(
        tabela,
        use_container_width=True,
        hide_index=True,
        key="editor_pendentes_dashboard",
        disabled=[
            col for col in tabela.columns
            if col not in ["Prioridade", "Concluir"]
        ],
        column_config={
            "Prioridade": st.column_config.SelectboxColumn(
                "Prioridade",
                options=["Alta", "Média", "Baixa"],
                required=True
            ),
            "Concluir": st.column_config.CheckboxColumn(
                "Concluir",
                help="Marque para concluir a localidade",
                default=False
            ),
            "__linha_original": None
        }
    )

    if st.button("Salvar alterações de pendentes"):
        indices_concluir = []

        for _, linha in tabela_editada.iterrows():
            indice_original = int(linha["__linha_original"])
            nova_prioridade = str(linha["Prioridade"]).strip()

            if nova_prioridade in ["Alta", "Média", "Baixa"]:
                df_pendentes.loc[indice_original,
                                 "Prioridade"] = nova_prioridade

            if bool(linha["Concluir"]):
                indices_concluir.append(indice_original)

        if indices_concluir:
            localidades_concluidas = df_pendentes.loc[
                indices_concluir,
                ["UF", "Cidade"]
            ].copy()

            df_concluidas = pd.concat(
                [df_concluidas, localidades_concluidas],
                ignore_index=True
            )

            df_concluidas = df_concluidas.drop_duplicates(
                subset=["UF", "Cidade"],
                keep="first"
            )

            df_pendentes = df_pendentes.drop(
                index=indices_concluir
            ).reset_index(drop=True)

        salvar_google_sheets(df_total, df_pendentes, df_concluidas, df_buscas)

        st.success("Alterações salvas com sucesso.")
        st.rerun()


with aba_relatorio:
    st.subheader("Relatório Detalhado das Localidades Pendentes")
    st.caption(
        "Consulte, filtre e edite o texto do relatório detalhado diretamente pelos cards."
    )

    col_rel1, col_rel2, col_rel3 = st.columns(3)

    with col_rel1:
        busca_relatorio = st.text_input("Pesquisar localidade no relatório")

    with col_rel2:
        uf_relatorio = st.selectbox(
            "Filtrar UF no relatório",
            ["Todas"] + sorted(df_pendentes["UF"].dropna().unique().tolist())
        )

    with col_rel3:
        prioridade_relatorio = st.selectbox(
            "Filtrar prioridade no relatório",
            ["Todas", "Alta", "Média", "Baixa"]
        )

    tabela_relatorio = df_pendentes.copy()
    tabela_relatorio["__linha_original"] = tabela_relatorio.index

    if busca_relatorio:
        tabela_relatorio = tabela_relatorio[
            tabela_relatorio["Cidade"].astype(str).str.contains(
                busca_relatorio,
                case=False,
                na=False
            )
        ]

    if uf_relatorio != "Todas":
        tabela_relatorio = tabela_relatorio[
            tabela_relatorio["UF"] == uf_relatorio
        ]

    if prioridade_relatorio != "Todas":
        tabela_relatorio = tabela_relatorio[
            tabela_relatorio["Prioridade"] == prioridade_relatorio
        ]

    st.write(f"Total filtrado: {len(tabela_relatorio)} localidade(s)")

    relatorios_editados = {}

    for _, linha in tabela_relatorio.iterrows():
        indice_original = int(linha["__linha_original"])
        cidade = linha.get("Cidade", "")
        uf = linha.get("UF", "")
        prioridade = linha.get("Prioridade", "")
        data = linha.get("Data da Solicitação", "")
        previsao = linha.get("Previsão", "")
        relatorio = linha.get("Relatório Detalhado", "")

        with st.container(border=True):
            st.markdown(f"### {cidade} - {uf}")
            st.markdown(f"**Prioridade:** {prioridade}")
            st.markdown(f"**Data da Solicitação:** {data}")
            st.markdown(f"**Previsão:** {previsao}")
            st.markdown("**Situação:**")

            novo_relatorio = st.text_area(
                label="Editar relatório",
                value=str(relatorio),
                key=f"relatorio_editavel_{indice_original}",
                height=120,
                label_visibility="collapsed"
            )

            relatorios_editados[indice_original] = novo_relatorio

    col_salvar_rel1, col_salvar_rel2 = st.columns([1, 3])

    with col_salvar_rel1:
        if st.button("Salvar alterações do relatório"):
            for indice_original, novo_relatorio in relatorios_editados.items():
                df_pendentes.loc[
                    indice_original,
                    "Relatório Detalhado"
                ] = str(novo_relatorio).strip()

            salvar_google_sheets(df_total, df_pendentes,
                                 df_concluidas, df_buscas)

            st.success("Relatório detalhado atualizado com sucesso.")
            st.rerun()

    with col_salvar_rel2:
        if st.button("Salvar textos automáticos no Google Sheets"):
            df_pendentes = preencher_relatorios_vazios(df_pendentes)
            salvar_google_sheets(df_total, df_pendentes,
                                 df_concluidas, df_buscas)
            st.success(
                "Relatórios vazios preenchidos e salvos no Google Sheets.")
            st.rerun()


with aba_buscas:
    st.subheader("Gestão de Buscas")
    st.caption(
        "Controle das localidades que Felipe e Diovane estão buscando, com status e observações."
    )

    col_b1, col_b2 = st.columns(2)
    col_b1.metric("Buscas Felipe", len(
        df_buscas[df_buscas["Agente"] == "Felipe"]))
    col_b2.metric("Buscas Diovane", len(
        df_buscas[df_buscas["Agente"] == "Diovane"]))

    st.divider()

    st.subheader("Adicionar Nova Busca")

    with st.expander("Cadastrar busca", expanded=False):
        busca_col1, busca_col2, busca_col3 = st.columns(3)

        with busca_col1:
            novo_agente = st.selectbox("Agente", AGENTES)
            nova_uf_busca = st.text_input("UF da busca", max_chars=2).upper()

        with busca_col2:
            nova_cidade_busca = st.text_input("Cidade da busca")
            nova_data_busca = st.date_input("Data da Busca")

        with busca_col3:
            novo_status_busca = st.selectbox("Status", STATUS_BUSCA)
            nova_observacao_busca = st.text_area("Observação", height=100)

        if st.button("Adicionar Busca"):
            if not nova_uf_busca or not nova_cidade_busca:
                st.warning("Preencha pelo menos UF e Cidade.")
            else:
                nova_linha_busca = {
                    "Agente": novo_agente,
                    "UF": nova_uf_busca.strip().upper(),
                    "Cidade": nova_cidade_busca.strip().title(),
                    "Data da Busca": nova_data_busca.strftime("%d/%m/%Y"),
                    "Status": novo_status_busca,
                    "Observação": nova_observacao_busca.strip()
                }

                df_buscas = pd.concat(
                    [df_buscas, pd.DataFrame([nova_linha_busca])],
                    ignore_index=True
                )

                salvar_google_sheets(df_total, df_pendentes,
                                     df_concluidas, df_buscas)

                st.success("Busca adicionada com sucesso.")
                st.rerun()

    st.divider()

    st.subheader("Tabela de Buscas")

    filtro_busca1, filtro_busca2, filtro_busca3 = st.columns(3)

    with filtro_busca1:
        filtro_agente = st.selectbox(
            "Filtrar por agente",
            ["Todos"] + AGENTES
        )

    with filtro_busca2:
        filtro_status = st.selectbox(
            "Filtrar por status",
            ["Todos"] + STATUS_BUSCA
        )

    with filtro_busca3:
        filtro_cidade_busca = st.text_input("Pesquisar cidade na gestão")

    tabela_buscas = df_buscas.copy()
    tabela_buscas["__linha_original"] = tabela_buscas.index
    tabela_buscas["Excluir"] = False

    if filtro_agente != "Todos":
        tabela_buscas = tabela_buscas[tabela_buscas["Agente"] == filtro_agente]

    if filtro_status != "Todos":
        tabela_buscas = tabela_buscas[tabela_buscas["Status"] == filtro_status]

    if filtro_cidade_busca:
        tabela_buscas = tabela_buscas[
            tabela_buscas["Cidade"].astype(str).str.contains(
                filtro_cidade_busca,
                case=False,
                na=False
            )
        ]

    if tabela_buscas.empty:
        st.info("Nenhuma busca encontrada para os filtros selecionados.")
    else:
        tabela_buscas_editada = st.data_editor(
            tabela_buscas[
                [
                    "Agente",
                    "UF",
                    "Cidade",
                    "Data da Busca",
                    "Status",
                    "Observação",
                    "Excluir",
                    "__linha_original"
                ]
            ],
            use_container_width=True,
            hide_index=True,
            key="editor_buscas",
            disabled=["UF", "Cidade", "Data da Busca", "__linha_original"],
            column_config={
                "Agente": st.column_config.SelectboxColumn(
                    "Agente",
                    options=AGENTES,
                    required=True
                ),
                "Status": st.column_config.SelectboxColumn(
                    "Status",
                    options=STATUS_BUSCA,
                    required=True
                ),
                "Observação": st.column_config.TextColumn(
                    "Observação"
                ),
                "Excluir": st.column_config.CheckboxColumn(
                    "Excluir",
                    help="Marque para remover esta busca",
                    default=False
                ),
                "__linha_original": None
            }
        )

        if st.button("Salvar Gestão de Buscas"):
            indices_excluir_busca = []

            for _, linha in tabela_buscas_editada.iterrows():
                indice_original = int(linha["__linha_original"])

                df_buscas.loc[indice_original, "Agente"] = str(
                    linha["Agente"]).strip()
                df_buscas.loc[indice_original, "Status"] = str(
                    linha["Status"]).strip()
                df_buscas.loc[indice_original, "Observação"] = str(
                    linha["Observação"]).strip()

                if bool(linha["Excluir"]):
                    indices_excluir_busca.append(indice_original)

            if indices_excluir_busca:
                df_buscas = df_buscas.drop(
                    index=indices_excluir_busca
                ).reset_index(drop=True)

            salvar_google_sheets(df_total, df_pendentes,
                                 df_concluidas, df_buscas)

            st.success("Gestão de buscas atualizada com sucesso.")
            st.rerun()


with aba_total:
    st.subheader("Localidades Total")
    st.caption(
        "Lista geral de localidades abertas, organizada em ordem alfabética.")

    filtro_total1, filtro_total2 = st.columns(2)

    with filtro_total1:
        busca_total = st.text_input("Pesquisar cidade no total")

    with filtro_total2:
        uf_total = st.selectbox(
            "Filtrar UF no total",
            ["Todas"] + sorted(df_total["UF"].dropna().unique().tolist())
        )

    tabela_total = df_total.copy()
    tabela_total["__linha_original"] = tabela_total.index
    tabela_total["Excluir"] = False

    if busca_total:
        tabela_total = tabela_total[
            tabela_total["Cidade"].astype(str).str.contains(
                busca_total,
                case=False,
                na=False
            )
        ]

    if uf_total != "Todas":
        tabela_total = tabela_total[tabela_total["UF"] == uf_total]

    tabela_total_editada = st.data_editor(
        tabela_total[["UF", "Cidade", "Excluir", "__linha_original"]],
        use_container_width=True,
        hide_index=True,
        key="editor_localidades_total",
        disabled=["UF", "Cidade", "__linha_original"],
        column_config={
            "Excluir": st.column_config.CheckboxColumn(
                "Excluir",
                help="Marque somente se a localidade foi cadastrada por engano",
                default=False
            ),
            "__linha_original": None
        }
    )

    confirmar_exclusao_total = st.checkbox(
        "Confirmo que desejo excluir as localidades marcadas do Total, Pendentes e Concluídas"
    )

    if st.button("Excluir localidades marcadas"):
        if not confirmar_exclusao_total:
            st.warning("Marque a confirmação antes de excluir.")
        else:
            linhas_para_excluir = tabela_total_editada[
                tabela_total_editada["Excluir"] == True
            ]

            if linhas_para_excluir.empty:
                st.warning("Nenhuma localidade foi marcada para exclusão.")
            else:
                indices_total_excluir = linhas_para_excluir[
                    "__linha_original"
                ].astype(int).tolist()

                localidades_excluir = df_total.loc[
                    indices_total_excluir,
                    ["UF", "Cidade"]
                ].copy()

                df_total = df_total.drop(
                    index=indices_total_excluir
                ).reset_index(drop=True)

                for _, localidade in localidades_excluir.iterrows():
                    uf_excluir = str(localidade["UF"]).upper().strip()
                    cidade_excluir = str(localidade["Cidade"]).upper().strip()

                    df_pendentes = df_pendentes[
                        ~(
                            (df_pendentes["UF"].astype(str).str.upper().str.strip() == uf_excluir) &
                            (df_pendentes["Cidade"].astype(
                                str).str.upper().str.strip() == cidade_excluir)
                        )
                    ].reset_index(drop=True)

                    df_concluidas = df_concluidas[
                        ~(
                            (df_concluidas["UF"].astype(str).str.upper().str.strip() == uf_excluir) &
                            (df_concluidas["Cidade"].astype(
                                str).str.upper().str.strip() == cidade_excluir)
                        )
                    ].reset_index(drop=True)

                salvar_google_sheets(df_total, df_pendentes,
                                     df_concluidas, df_buscas)

                st.success("Localidade(s) excluída(s) com sucesso.")
                st.rerun()


with aba_concluidas:
    st.subheader("Localidades Concluídas")
    st.caption(
        "Lista de localidades concluídas. Também é possível retornar uma localidade para pendentes."
    )

    filtro_conc1, filtro_conc2 = st.columns(2)

    with filtro_conc1:
        busca_concluida = st.text_input("Pesquisar cidade concluída")

    with filtro_conc2:
        uf_concluida = st.selectbox(
            "Filtrar UF concluída",
            ["Todas"] + sorted(df_concluidas["UF"].dropna().unique().tolist())
        )

    tabela_concluidas = df_concluidas.copy()
    tabela_concluidas["__linha_original"] = tabela_concluidas.index
    tabela_concluidas["Retornar para Pendentes"] = False

    if busca_concluida:
        tabela_concluidas = tabela_concluidas[
            tabela_concluidas["Cidade"].astype(str).str.contains(
                busca_concluida,
                case=False,
                na=False
            )
        ]

    if uf_concluida != "Todas":
        tabela_concluidas = tabela_concluidas[
            tabela_concluidas["UF"] == uf_concluida
        ]

    if tabela_concluidas.empty:
        st.info("Nenhuma localidade concluída encontrada.")
    else:
        tabela_concluidas_editada = st.data_editor(
            tabela_concluidas[
                ["UF", "Cidade", "Retornar para Pendentes", "__linha_original"]
            ],
            use_container_width=True,
            hide_index=True,
            key="editor_concluidas",
            disabled=["UF", "Cidade", "__linha_original"],
            column_config={
                "Retornar para Pendentes": st.column_config.CheckboxColumn(
                    "Retornar para Pendentes",
                    help="Marque para retornar a localidade para pendentes",
                    default=False
                ),
                "__linha_original": None
            }
        )

        confirmar_retorno = st.checkbox(
            "Confirmo que desejo retornar as localidades marcadas para Pendentes"
        )

        if st.button("Retornar localidades marcadas"):
            if not confirmar_retorno:
                st.warning("Marque a confirmação antes de retornar.")
            else:
                linhas_retorno = tabela_concluidas_editada[
                    tabela_concluidas_editada["Retornar para Pendentes"] == True
                ]

                if linhas_retorno.empty:
                    st.warning("Nenhuma localidade foi marcada para retornar.")
                else:
                    indices_retorno = linhas_retorno[
                        "__linha_original"
                    ].astype(int).tolist()

                    localidades_retorno = df_concluidas.loc[
                        indices_retorno,
                        ["UF", "Cidade"]
                    ].copy()

                    hoje = datetime.now().strftime("%d/%m/%Y")

                    novas_pendentes = []

                    for indice, linha in localidades_retorno.iterrows():
                        novas_pendentes.append({
                            "UF": str(linha["UF"]).strip().upper(),
                            "Cidade": str(linha["Cidade"]).strip(),
                            "Data da Solicitação": hoje,
                            "Previsão": "",
                            "Prioridade": "Alta",
                            "Relatório Detalhado": RELATORIOS_PADRAO[
                                len(df_pendentes) % len(RELATORIOS_PADRAO)
                            ]
                        })

                    df_pendentes = pd.concat(
                        [df_pendentes, pd.DataFrame(novas_pendentes)],
                        ignore_index=True
                    )

                    df_pendentes = df_pendentes.drop_duplicates(
                        subset=["UF", "Cidade"],
                        keep="last"
                    ).reset_index(drop=True)

                    df_concluidas = df_concluidas.drop(
                        index=indices_retorno
                    ).reset_index(drop=True)

                    salvar_google_sheets(
                        df_total, df_pendentes, df_concluidas, df_buscas)

                    st.success(
                        "Localidade(s) retornada(s) para Pendentes com sucesso.")
                    st.rerun()


st.divider()
st.subheader("Exportação de Relatórios")

col_down1, col_down2, col_down3, col_down4 = st.columns(4)

with col_down1:
    st.download_button(
        label="Baixar pendentes em CSV",
        data=df_pendentes.to_csv(index=False).encode("utf-8-sig"),
        file_name="localidades_pendentes.csv",
        mime="text/csv"
    )

with col_down2:
    st.download_button(
        label="Baixar concluídas em CSV",
        data=df_concluidas.to_csv(index=False).encode("utf-8-sig"),
        file_name="localidades_concluidas.csv",
        mime="text/csv"
    )

with col_down3:
    st.download_button(
        label="Baixar buscas em CSV",
        data=df_buscas.to_csv(index=False).encode("utf-8-sig"),
        file_name="buscas_por_agente.csv",
        mime="text/csv"
    )

with col_down4:
    st.download_button(
        label="Baixar relatório detalhado em Excel",
        data=gerar_excel_relatorio(df_pendentes),
        file_name="relatorio_detalhado.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
